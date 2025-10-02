"""
数据统计和分析服务
提供用户统计、事件统计、处理效率分析、热点区域分析等功能
"""
import asyncio
import json
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any
from uuid import UUID

import pandas as pd
from sqlalchemy import and_, func, select, text
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import AsyncSessionLocal
from app.models.event import Event, EventStatus, EventTimeline, EventPriority
from app.models.user import User
from app.core.redis_client import redis_client

logger = logging.getLogger(__name__)


class StatisticsService:
    """统计分析服务"""
    
    def __init__(self):
        self.cache_prefix = "stats:"
        self.cache_ttl = 3600  # 1小时缓存
    
    async def get_user_statistics(
        self, 
        user_id: Optional[UUID] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> Dict[str, Any]:
        """
        获取用户统计数据
        
        Args:
            user_id: 用户ID，如果为None则统计所有用户
            start_date: 开始日期
            end_date: 结束日期
            
        Returns:
            用户统计数据字典
        """
        cache_key = f"{self.cache_prefix}user:{user_id}:{start_date}:{end_date}"
        
        # 尝试从缓存获取
        cached_result = await redis_client.get(cache_key)
        if cached_result:
            return json.loads(cached_result)
        
        async with AsyncSessionLocal() as db:
            # 构建查询条件
            conditions = []
            if user_id:
                conditions.append(Event.user_id == user_id)
            if start_date:
                conditions.append(Event.created_at >= start_date)
            if end_date:
                conditions.append(Event.created_at <= end_date)
            
            # 查询用户事件统计
            query = select(
                func.count(Event.id).label('total_events'),
                func.count(func.nullif(Event.status == EventStatus.COMPLETED, False)).label('completed_events'),
                func.count(func.nullif(Event.status == EventStatus.PENDING, False)).label('pending_events'),
                func.count(func.nullif(Event.status == EventStatus.PROCESSING, False)).label('processing_events'),
                func.count(func.nullif(Event.status == EventStatus.REJECTED, False)).label('rejected_events'),
                func.avg(Event.confidence).label('avg_confidence')
            )
            
            if conditions:
                query = query.where(and_(*conditions))
            
            result = await db.execute(query)
            stats = result.first()
            
            # 计算完成率
            total_events = stats.total_events or 0
            completed_events = stats.completed_events or 0
            completion_rate = (completed_events / total_events * 100) if total_events > 0 else 0
            
            # 查询事件类型分布
            type_query = select(
                Event.event_type,
                func.count(Event.id).label('count')
            ).group_by(Event.event_type)
            
            if conditions:
                type_query = type_query.where(and_(*conditions))
            
            type_result = await db.execute(type_query)
            event_types = [
                {"type": row.event_type, "count": row.count}
                for row in type_result
            ]
            
            # 查询优先级分布
            priority_query = select(
                Event.priority,
                func.count(Event.id).label('count')
            ).group_by(Event.priority)
            
            if conditions:
                priority_query = priority_query.where(and_(*conditions))
            
            priority_result = await db.execute(priority_query)
            priorities = [
                {"priority": row.priority.value if row.priority else None, "count": row.count}
                for row in priority_result
            ]
            
            statistics = {
                "total_events": total_events,
                "completed_events": completed_events,
                "pending_events": stats.pending_events or 0,
                "processing_events": stats.processing_events or 0,
                "rejected_events": stats.rejected_events or 0,
                "completion_rate": round(completion_rate, 2),
                "avg_confidence": round(float(stats.avg_confidence or 0), 2),
                "event_types": event_types,
                "priorities": priorities,
                "generated_at": datetime.utcnow().isoformat()
            }
            
            # 缓存结果
            await redis_client.setex(
                cache_key, 
                self.cache_ttl, 
                json.dumps(statistics, default=str)
            )
            
            return statistics
    
    async def get_event_statistics(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        group_by: str = "day"  # day, week, month
    ) -> Dict[str, Any]:
        """
        获取事件统计分析数据
        
        Args:
            start_date: 开始日期
            end_date: 结束日期
            group_by: 分组方式 (day, week, month)
            
        Returns:
            事件统计数据
        """
        cache_key = f"{self.cache_prefix}events:{start_date}:{end_date}:{group_by}"
        
        # 尝试从缓存获取
        cached_result = await redis_client.get(cache_key)
        if cached_result:
            return json.loads(cached_result)
        
        async with AsyncSessionLocal() as db:
            # 设置默认时间范围
            if not end_date:
                end_date = datetime.utcnow()
            if not start_date:
                start_date = end_date - timedelta(days=30)
            
            # 根据分组方式构建时间分组函数
            if group_by == "day":
                time_group = func.date_trunc('day', Event.created_at)
            elif group_by == "week":
                time_group = func.date_trunc('week', Event.created_at)
            elif group_by == "month":
                time_group = func.date_trunc('month', Event.created_at)
            else:
                time_group = func.date_trunc('day', Event.created_at)
            
            # 查询时间序列统计
            time_series_query = select(
                time_group.label('time_period'),
                func.count(Event.id).label('total_count'),
                func.count(func.nullif(Event.status == EventStatus.COMPLETED, False)).label('completed_count'),
                func.count(func.nullif(Event.status == EventStatus.PENDING, False)).label('pending_count'),
                func.count(func.nullif(Event.status == EventStatus.PROCESSING, False)).label('processing_count')
            ).where(
                and_(
                    Event.created_at >= start_date,
                    Event.created_at <= end_date
                )
            ).group_by(time_group).order_by(time_group)
            
            time_series_result = await db.execute(time_series_query)
            time_series = [
                {
                    "period": row.time_period.isoformat() if row.time_period else None,
                    "total": row.total_count,
                    "completed": row.completed_count,
                    "pending": row.pending_count,
                    "processing": row.processing_count
                }
                for row in time_series_result
            ]
            
            # 查询事件类型统计
            type_stats_query = select(
                Event.event_type,
                func.count(Event.id).label('total_count'),
                func.count(func.nullif(Event.status == EventStatus.COMPLETED, False)).label('completed_count'),
                func.avg(Event.confidence).label('avg_confidence')
            ).where(
                and_(
                    Event.created_at >= start_date,
                    Event.created_at <= end_date
                )
            ).group_by(Event.event_type).order_by(func.count(Event.id).desc())
            
            type_stats_result = await db.execute(type_stats_query)
            type_statistics = [
                {
                    "event_type": row.event_type,
                    "total_count": row.total_count,
                    "completed_count": row.completed_count,
                    "completion_rate": round((row.completed_count / row.total_count * 100) if row.total_count > 0 else 0, 2),
                    "avg_confidence": round(float(row.avg_confidence or 0), 2)
                }
                for row in type_stats_result
            ]
            
            # 查询状态分布
            status_query = select(
                Event.status,
                func.count(Event.id).label('count')
            ).where(
                and_(
                    Event.created_at >= start_date,
                    Event.created_at <= end_date
                )
            ).group_by(Event.status)
            
            status_result = await db.execute(status_query)
            status_distribution = [
                {
                    "status": row.status.value if row.status else None,
                    "count": row.count
                }
                for row in status_result
            ]
            
            statistics = {
                "time_series": time_series,
                "type_statistics": type_statistics,
                "status_distribution": status_distribution,
                "period": {
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat(),
                    "group_by": group_by
                },
                "generated_at": datetime.utcnow().isoformat()
            }
            
            # 缓存结果
            await redis_client.setex(
                cache_key,
                self.cache_ttl,
                json.dumps(statistics, default=str)
            )
            
            return statistics    
    async def get_processing_efficiency_analysis(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> Dict[str, Any]:
        """
        获取处理效率分析数据
        
        Args:
            start_date: 开始日期
            end_date: 结束日期
            
        Returns:
            处理效率分析数据
        """
        cache_key = f"{self.cache_prefix}efficiency:{start_date}:{end_date}"
        
        # 尝试从缓存获取
        cached_result = await redis_client.get(cache_key)
        if cached_result:
            return json.loads(cached_result)
        
        async with AsyncSessionLocal() as db:
            # 设置默认时间范围
            if not end_date:
                end_date = datetime.utcnow()
            if not start_date:
                start_date = end_date - timedelta(days=30)
            
            # 查询已完成事件的处理时间
            completed_events_query = select(
                Event.id,
                Event.event_type,
                Event.priority,
                Event.created_at,
                func.max(EventTimeline.created_at).label('completed_at')
            ).select_from(
                Event.join(EventTimeline, Event.id == EventTimeline.event_id)
            ).where(
                and_(
                    Event.status == EventStatus.COMPLETED,
                    Event.created_at >= start_date,
                    Event.created_at <= end_date,
                    EventTimeline.status == EventStatus.COMPLETED
                )
            ).group_by(Event.id, Event.event_type, Event.priority, Event.created_at)
            
            completed_events_result = await db.execute(completed_events_query)
            completed_events = completed_events_result.all()
            
            # 计算处理时间
            processing_times = []
            for event in completed_events:
                if event.completed_at and event.created_at:
                    processing_time = (event.completed_at - event.created_at).total_seconds() / 3600  # 转换为小时
                    processing_times.append({
                        "event_id": str(event.id),
                        "event_type": event.event_type,
                        "priority": event.priority.value if event.priority else None,
                        "processing_time_hours": round(processing_time, 2)
                    })
            
            # 计算总体效率指标
            if processing_times:
                times = [pt["processing_time_hours"] for pt in processing_times]
                avg_processing_time = sum(times) / len(times)
                min_processing_time = min(times)
                max_processing_time = max(times)
                
                # 计算按事件类型的平均处理时间
                type_efficiency = {}
                for pt in processing_times:
                    event_type = pt["event_type"]
                    if event_type not in type_efficiency:
                        type_efficiency[event_type] = []
                    type_efficiency[event_type].append(pt["processing_time_hours"])
                
                type_avg_times = [
                    {
                        "event_type": event_type,
                        "avg_processing_time": round(sum(times) / len(times), 2),
                        "event_count": len(times)
                    }
                    for event_type, times in type_efficiency.items()
                ]
                
                # 计算按优先级的平均处理时间
                priority_efficiency = {}
                for pt in processing_times:
                    priority = pt["priority"] or "unknown"
                    if priority not in priority_efficiency:
                        priority_efficiency[priority] = []
                    priority_efficiency[priority].append(pt["processing_time_hours"])
                
                priority_avg_times = [
                    {
                        "priority": priority,
                        "avg_processing_time": round(sum(times) / len(times), 2),
                        "event_count": len(times)
                    }
                    for priority, times in priority_efficiency.items()
                ]
            else:
                avg_processing_time = 0
                min_processing_time = 0
                max_processing_time = 0
                type_avg_times = []
                priority_avg_times = []
            
            # 查询响应速度（从创建到开始处理的时间）
            response_time_query = select(
                Event.id,
                Event.event_type,
                Event.created_at,
                func.min(EventTimeline.created_at).label('first_response_at')
            ).select_from(
                Event.join(EventTimeline, Event.id == EventTimeline.event_id)
            ).where(
                and_(
                    Event.created_at >= start_date,
                    Event.created_at <= end_date,
                    EventTimeline.status == EventStatus.PROCESSING
                )
            ).group_by(Event.id, Event.event_type, Event.created_at)
            
            response_time_result = await db.execute(response_time_query)
            response_times = []
            
            for event in response_time_result:
                if event.first_response_at and event.created_at:
                    response_time = (event.first_response_at - event.created_at).total_seconds() / 3600  # 转换为小时
                    response_times.append({
                        "event_id": str(event.id),
                        "event_type": event.event_type,
                        "response_time_hours": round(response_time, 2)
                    })
            
            # 计算平均响应时间
            if response_times:
                times = [rt["response_time_hours"] for rt in response_times]
                avg_response_time = sum(times) / len(times)
            else:
                avg_response_time = 0
            
            efficiency_analysis = {
                "processing_efficiency": {
                    "avg_processing_time_hours": round(avg_processing_time, 2),
                    "min_processing_time_hours": round(min_processing_time, 2),
                    "max_processing_time_hours": round(max_processing_time, 2),
                    "completed_events_count": len(processing_times),
                    "type_efficiency": type_avg_times,
                    "priority_efficiency": priority_avg_times
                },
                "response_efficiency": {
                    "avg_response_time_hours": round(avg_response_time, 2),
                    "responded_events_count": len(response_times)
                },
                "period": {
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat()
                },
                "generated_at": datetime.utcnow().isoformat()
            }
            
            # 缓存结果
            await redis_client.setex(
                cache_key,
                self.cache_ttl,
                json.dumps(efficiency_analysis, default=str)
            )
            
            return efficiency_analysis
    
    async def get_hotspot_area_analysis(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        radius_km: float = 1.0  # 热点区域半径（公里）
    ) -> Dict[str, Any]:
        """
        获取热点区域分析数据
        
        Args:
            start_date: 开始日期
            end_date: 结束日期
            radius_km: 热点区域半径（公里）
            
        Returns:
            热点区域分析数据
        """
        cache_key = f"{self.cache_prefix}hotspots:{start_date}:{end_date}:{radius_km}"
        
        # 尝试从缓存获取
        cached_result = await redis_client.get(cache_key)
        if cached_result:
            return json.loads(cached_result)
        
        async with AsyncSessionLocal() as db:
            # 设置默认时间范围
            if not end_date:
                end_date = datetime.utcnow()
            if not start_date:
                start_date = end_date - timedelta(days=30)
            
            # 查询所有事件的地理位置
            events_query = select(
                Event.id,
                Event.event_type,
                Event.location_lat,
                Event.location_lng,
                Event.location_address,
                Event.created_at
            ).where(
                and_(
                    Event.created_at >= start_date,
                    Event.created_at <= end_date,
                    Event.location_lat.isnot(None),
                    Event.location_lng.isnot(None)
                )
            )
            
            events_result = await db.execute(events_query)
            events = events_result.all()
            
            if not events:
                return {
                    "hotspots": [],
                    "area_statistics": [],
                    "period": {
                        "start_date": start_date.isoformat(),
                        "end_date": end_date.isoformat()
                    },
                    "generated_at": datetime.utcnow().isoformat()
                }
            
            # 使用简单的网格聚类算法识别热点区域
            # 将地理坐标按网格分组
            grid_size = radius_km / 111.0  # 大约1公里对应的纬度差
            hotspot_grid = {}
            
            for event in events:
                # 计算网格坐标
                grid_lat = round(event.location_lat / grid_size) * grid_size
                grid_lng = round(event.location_lng / grid_size) * grid_size
                grid_key = f"{grid_lat},{grid_lng}"
                
                if grid_key not in hotspot_grid:
                    hotspot_grid[grid_key] = {
                        "center_lat": grid_lat,
                        "center_lng": grid_lng,
                        "events": [],
                        "event_types": {}
                    }
                
                hotspot_grid[grid_key]["events"].append({
                    "id": str(event.id),
                    "event_type": event.event_type,
                    "lat": event.location_lat,
                    "lng": event.location_lng,
                    "address": event.location_address,
                    "created_at": event.created_at.isoformat()
                })
                
                # 统计事件类型
                event_type = event.event_type
                if event_type not in hotspot_grid[grid_key]["event_types"]:
                    hotspot_grid[grid_key]["event_types"][event_type] = 0
                hotspot_grid[grid_key]["event_types"][event_type] += 1
            
            # 筛选热点区域（事件数量大于等于3的区域）
            hotspots = []
            for grid_key, grid_data in hotspot_grid.items():
                event_count = len(grid_data["events"])
                if event_count >= 3:  # 至少3个事件才算热点
                    # 计算实际中心点（所有事件的平均位置）
                    avg_lat = sum(e["lat"] for e in grid_data["events"]) / event_count
                    avg_lng = sum(e["lng"] for e in grid_data["events"]) / event_count
                    
                    # 获取最常见的地址
                    addresses = [e["address"] for e in grid_data["events"] if e["address"]]
                    most_common_address = max(set(addresses), key=addresses.count) if addresses else "未知地址"
                    
                    hotspots.append({
                        "center_lat": round(avg_lat, 6),
                        "center_lng": round(avg_lng, 6),
                        "address": most_common_address,
                        "event_count": event_count,
                        "event_types": grid_data["event_types"],
                        "dominant_type": max(grid_data["event_types"], key=grid_data["event_types"].get),
                        "events": grid_data["events"]
                    })
            
            # 按事件数量排序
            hotspots.sort(key=lambda x: x["event_count"], reverse=True)
            
            # 计算区域统计（按地址前缀分组）
            area_stats = {}
            for event in events:
                if event.location_address:
                    # 提取地址的前两级（如：浙江省杭州市）
                    address_parts = event.location_address.split()
                    if len(address_parts) >= 2:
                        area_key = " ".join(address_parts[:2])
                    else:
                        area_key = address_parts[0] if address_parts else "未知区域"
                    
                    if area_key not in area_stats:
                        area_stats[area_key] = {
                            "area_name": area_key,
                            "event_count": 0,
                            "event_types": {}
                        }
                    
                    area_stats[area_key]["event_count"] += 1
                    event_type = event.event_type
                    if event_type not in area_stats[area_key]["event_types"]:
                        area_stats[area_key]["event_types"][event_type] = 0
                    area_stats[area_key]["event_types"][event_type] += 1
            
            area_statistics = list(area_stats.values())
            area_statistics.sort(key=lambda x: x["event_count"], reverse=True)
            
            hotspot_analysis = {
                "hotspots": hotspots[:20],  # 返回前20个热点区域
                "area_statistics": area_statistics[:10],  # 返回前10个区域统计
                "total_events": len(events),
                "total_hotspots": len(hotspots),
                "parameters": {
                    "radius_km": radius_km,
                    "min_events_for_hotspot": 3
                },
                "period": {
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat()
                },
                "generated_at": datetime.utcnow().isoformat()
            }
            
            # 缓存结果
            await redis_client.setex(
                cache_key,
                self.cache_ttl,
                json.dumps(hotspot_analysis, default=str)
            )
            
            return hotspot_analysis    
    async def export_statistics_to_excel(
        self,
        export_type: str = "user_stats",  # user_stats, event_stats, efficiency, hotspots
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        user_id: Optional[UUID] = None
    ) -> bytes:
        """
        导出统计数据到Excel格式
        
        Args:
            export_type: 导出类型
            start_date: 开始日期
            end_date: 结束日期
            user_id: 用户ID（仅用于用户统计）
            
        Returns:
            Excel文件的字节数据
        """
        try:
            # 根据导出类型获取数据
            if export_type == "user_stats":
                data = await self.get_user_statistics(user_id, start_date, end_date)
                return self._create_user_stats_excel(data)
            elif export_type == "event_stats":
                data = await self.get_event_statistics(start_date, end_date)
                return self._create_event_stats_excel(data)
            elif export_type == "efficiency":
                data = await self.get_processing_efficiency_analysis(start_date, end_date)
                return self._create_efficiency_excel(data)
            elif export_type == "hotspots":
                data = await self.get_hotspot_area_analysis(start_date, end_date)
                return self._create_hotspots_excel(data)
            else:
                raise ValueError(f"不支持的导出类型: {export_type}")
                
        except Exception as e:
            logger.error(f"导出Excel失败: {str(e)}")
            raise
    
    async def export_statistics_to_csv(
        self,
        export_type: str = "user_stats",
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        user_id: Optional[UUID] = None
    ) -> str:
        """
        导出统计数据到CSV格式
        
        Args:
            export_type: 导出类型
            start_date: 开始日期
            end_date: 结束日期
            user_id: 用户ID（仅用于用户统计）
            
        Returns:
            CSV格式的字符串数据
        """
        try:
            # 根据导出类型获取数据
            if export_type == "user_stats":
                data = await self.get_user_statistics(user_id, start_date, end_date)
                return self._create_user_stats_csv(data)
            elif export_type == "event_stats":
                data = await self.get_event_statistics(start_date, end_date)
                return self._create_event_stats_csv(data)
            elif export_type == "efficiency":
                data = await self.get_processing_efficiency_analysis(start_date, end_date)
                return self._create_efficiency_csv(data)
            elif export_type == "hotspots":
                data = await self.get_hotspot_area_analysis(start_date, end_date)
                return self._create_hotspots_csv(data)
            else:
                raise ValueError(f"不支持的导出类型: {export_type}")
                
        except Exception as e:
            logger.error(f"导出CSV失败: {str(e)}")
            raise
    
    def _create_user_stats_excel(self, data: Dict[str, Any]) -> bytes:
        """创建用户统计Excel文件"""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "用户统计"
        
        # 设置标题
        ws['A1'] = "用户统计报告"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')
        
        # 基础统计
        row = 3
        ws[f'A{row}'] = "总事件数"
        ws[f'B{row}'] = data['total_events']
        row += 1
        ws[f'A{row}'] = "已完成事件数"
        ws[f'B{row}'] = data['completed_events']
        row += 1
        ws[f'A{row}'] = "完成率(%)"
        ws[f'B{row}'] = data['completion_rate']
        row += 1
        ws[f'A{row}'] = "平均置信度"
        ws[f'B{row}'] = data['avg_confidence']
        
        # 事件类型分布
        row += 2
        ws[f'A{row}'] = "事件类型分布"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = "事件类型"
        ws[f'B{row}'] = "数量"
        
        for event_type in data['event_types']:
            row += 1
            ws[f'A{row}'] = event_type['type']
            ws[f'B{row}'] = event_type['count']
        
        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def _create_user_stats_csv(self, data: Dict[str, Any]) -> str:
        """创建用户统计CSV文件"""
        import io
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 写入基础统计
        writer.writerow(["用户统计报告"])
        writer.writerow([])
        writer.writerow(["指标", "数值"])
        writer.writerow(["总事件数", data['total_events']])
        writer.writerow(["已完成事件数", data['completed_events']])
        writer.writerow(["完成率(%)", data['completion_rate']])
        writer.writerow(["平均置信度", data['avg_confidence']])
        
        # 写入事件类型分布
        writer.writerow([])
        writer.writerow(["事件类型分布"])
        writer.writerow(["事件类型", "数量"])
        for event_type in data['event_types']:
            writer.writerow([event_type['type'], event_type['count']])
        
        return output.getvalue()
    
    def _create_event_stats_excel(self, data: Dict[str, Any]) -> bytes:
        """创建事件统计Excel文件"""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        wb = Workbook()
        
        # 时间序列数据
        ws1 = wb.active
        ws1.title = "时间序列"
        ws1['A1'] = "时间序列统计"
        ws1['A1'].font = Font(size=14, bold=True)
        
        ws1['A3'] = "时间"
        ws1['B3'] = "总数"
        ws1['C3'] = "已完成"
        ws1['D3'] = "待处理"
        ws1['E3'] = "处理中"
        
        row = 4
        for item in data['time_series']:
            ws1[f'A{row}'] = item['period']
            ws1[f'B{row}'] = item['total']
            ws1[f'C{row}'] = item['completed']
            ws1[f'D{row}'] = item['pending']
            ws1[f'E{row}'] = item['processing']
            row += 1
        
        # 事件类型统计
        ws2 = wb.create_sheet("事件类型统计")
        ws2['A1'] = "事件类型统计"
        ws2['A1'].font = Font(size=14, bold=True)
        
        ws2['A3'] = "事件类型"
        ws2['B3'] = "总数"
        ws2['C3'] = "已完成"
        ws2['D3'] = "完成率(%)"
        ws2['E3'] = "平均置信度"
        
        row = 4
        for item in data['type_statistics']:
            ws2[f'A{row}'] = item['event_type']
            ws2[f'B{row}'] = item['total_count']
            ws2[f'C{row}'] = item['completed_count']
            ws2[f'D{row}'] = item['completion_rate']
            ws2[f'E{row}'] = item['avg_confidence']
            row += 1
        
        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def _create_event_stats_csv(self, data: Dict[str, Any]) -> str:
        """创建事件统计CSV文件"""
        import io
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 时间序列数据
        writer.writerow(["时间序列统计"])
        writer.writerow(["时间", "总数", "已完成", "待处理", "处理中"])
        for item in data['time_series']:
            writer.writerow([item['period'], item['total'], item['completed'], item['pending'], item['processing']])
        
        writer.writerow([])
        
        # 事件类型统计
        writer.writerow(["事件类型统计"])
        writer.writerow(["事件类型", "总数", "已完成", "完成率(%)", "平均置信度"])
        for item in data['type_statistics']:
            writer.writerow([item['event_type'], item['total_count'], item['completed_count'], item['completion_rate'], item['avg_confidence']])
        
        return output.getvalue()
    
    def _create_efficiency_excel(self, data: Dict[str, Any]) -> bytes:
        """创建效率分析Excel文件"""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font
        
        wb = Workbook()
        ws = wb.active
        ws.title = "效率分析"
        
        ws['A1'] = "处理效率分析报告"
        ws['A1'].font = Font(size=16, bold=True)
        
        # 处理效率数据
        row = 3
        ws[f'A{row}'] = "平均处理时间(小时)"
        ws[f'B{row}'] = data['processing_efficiency']['avg_processing_time_hours']
        row += 1
        ws[f'A{row}'] = "最短处理时间(小时)"
        ws[f'B{row}'] = data['processing_efficiency']['min_processing_time_hours']
        row += 1
        ws[f'A{row}'] = "最长处理时间(小时)"
        ws[f'B{row}'] = data['processing_efficiency']['max_processing_time_hours']
        row += 1
        ws[f'A{row}'] = "平均响应时间(小时)"
        ws[f'B{row}'] = data['response_efficiency']['avg_response_time_hours']
        
        # 按类型的效率
        row += 2
        ws[f'A{row}'] = "按事件类型的处理效率"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = "事件类型"
        ws[f'B{row}'] = "平均处理时间(小时)"
        ws[f'C{row}'] = "事件数量"
        
        for item in data['processing_efficiency']['type_efficiency']:
            row += 1
            ws[f'A{row}'] = item['event_type']
            ws[f'B{row}'] = item['avg_processing_time']
            ws[f'C{row}'] = item['event_count']
        
        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def _create_efficiency_csv(self, data: Dict[str, Any]) -> str:
        """创建效率分析CSV文件"""
        import io
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow(["处理效率分析报告"])
        writer.writerow([])
        writer.writerow(["指标", "数值"])
        writer.writerow(["平均处理时间(小时)", data['processing_efficiency']['avg_processing_time_hours']])
        writer.writerow(["最短处理时间(小时)", data['processing_efficiency']['min_processing_time_hours']])
        writer.writerow(["最长处理时间(小时)", data['processing_efficiency']['max_processing_time_hours']])
        writer.writerow(["平均响应时间(小时)", data['response_efficiency']['avg_response_time_hours']])
        
        writer.writerow([])
        writer.writerow(["按事件类型的处理效率"])
        writer.writerow(["事件类型", "平均处理时间(小时)", "事件数量"])
        for item in data['processing_efficiency']['type_efficiency']:
            writer.writerow([item['event_type'], item['avg_processing_time'], item['event_count']])
        
        return output.getvalue()
    
    def _create_hotspots_excel(self, data: Dict[str, Any]) -> bytes:
        """创建热点区域Excel文件"""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font
        
        wb = Workbook()
        ws = wb.active
        ws.title = "热点区域"
        
        ws['A1'] = "热点区域分析报告"
        ws['A1'].font = Font(size=16, bold=True)
        
        # 热点区域数据
        row = 3
        ws[f'A{row}'] = "中心纬度"
        ws[f'B{row}'] = "中心经度"
        ws[f'C{row}'] = "地址"
        ws[f'D{row}'] = "事件数量"
        ws[f'E{row}'] = "主要事件类型"
        
        for hotspot in data['hotspots']:
            row += 1
            ws[f'A{row}'] = hotspot['center_lat']
            ws[f'B{row}'] = hotspot['center_lng']
            ws[f'C{row}'] = hotspot['address']
            ws[f'D{row}'] = hotspot['event_count']
            ws[f'E{row}'] = hotspot['dominant_type']
        
        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def _create_hotspots_csv(self, data: Dict[str, Any]) -> str:
        """创建热点区域CSV文件"""
        import io
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow(["热点区域分析报告"])
        writer.writerow([])
        writer.writerow(["中心纬度", "中心经度", "地址", "事件数量", "主要事件类型"])
        for hotspot in data['hotspots']:
            writer.writerow([hotspot['center_lat'], hotspot['center_lng'], hotspot['address'], hotspot['event_count'], hotspot['dominant_type']])
        
        return output.getvalue()
    
    async def refresh_all_caches(self):
        """刷新所有统计数据缓存"""
        try:
            # 获取所有缓存键
            cache_keys = await redis_client.keys(f"{self.cache_prefix}*")
            
            if cache_keys:
                # 删除所有统计缓存
                await redis_client.delete(*cache_keys)
                logger.info(f"已清除 {len(cache_keys)} 个统计缓存")
            
            # 预热常用统计数据
            await self._warm_up_caches()
            
        except Exception as e:
            logger.error(f"刷新缓存失败: {str(e)}")
            raise
    
    async def _warm_up_caches(self):
        """预热缓存 - 生成常用的统计数据"""
        try:
            # 预热用户统计（最近30天）
            end_date = datetime.utcnow()
            start_date = end_date - timedelta(days=30)
            
            # 并发执行多个统计查询来预热缓存
            tasks = [
                self.get_user_statistics(start_date=start_date, end_date=end_date),
                self.get_event_statistics(start_date=start_date, end_date=end_date, group_by="day"),
                self.get_processing_efficiency_analysis(start_date=start_date, end_date=end_date),
                self.get_hotspot_area_analysis(start_date=start_date, end_date=end_date)
            ]
            
            await asyncio.gather(*tasks, return_exceptions=True)
            logger.info("统计缓存预热完成")
            
        except Exception as e:
            logger.error(f"缓存预热失败: {str(e)}")


# 创建全局统计服务实例
statistics_service = StatisticsService()